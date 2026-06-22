# extract_shear_data.py
# Purpose: Extract crack width measurements from shear01 filtered data
# Input: zijlwegdwars01.xls (converted to xlsx)
# Output: Verified crack width values at 5 load steps
# Author: Neguse Solomon Mekonnen - Dissertation 2025/2026

import openpyxl
import pandas as pd
import numpy as np

# Load the converted xlsx file
wb = openpyxl.load_workbook('zijlwegdwars01.xlsx', read_only=True, data_only=True)
ws = wb[wb.sheetnames[0]]

data = []
headers = None
for i, row in enumerate(ws.iter_rows(values_only=True)):
    if i == 0:
        headers = list(row)
    else:
        data.append(row)
wb.close()

df = pd.DataFrame(data, columns=headers)

# Total force = absolute sum of 4 jacks
df['F_total'] = abs(df['F1_Filtered'] + df['F2_Filtered'] + 
                    df['F3_Filtered'] + df['F4_Filtered'])

# Zero reference
zero_ref = df[df['F_total'] < 10]
ref14 = zero_ref['LVDT14_Filtered'].mean()
ref15 = zero_ref['LVDT15_Filtered'].mean()
ref16 = zero_ref['LVDT16_Filtered'].mean()

print(f"Zero references: LVDT14={ref14:.6f}, LVDT15={ref15:.6f}, LVDT16={ref16:.6f}")
print()

# Five load steps
load_steps = [400, 800, 1082, 1220, 1341]

print(f"{'Step':>5} {'F(kN)':>10} {'CW14(mm)':>12} {'CW15(mm)':>12} "
      f"{'CW16(mm)':>12} {'w_max,w(mm)':>14}")
print('-' * 75)

w_max_list = []
F_list = []
for target in load_steps:
    mask = (df['F_total'] >= target - 50) & (df['F_total'] <= target + 50)
    s = df[mask]
    if len(s) > 0:
        F = s['F_total'].mean()
        # Crack widths already in mm (verified against Lantsoght 2017 Fig 12c)
        cw14 = s['LVDT14_Filtered'].mean() - ref14
        cw15 = s['LVDT15_Filtered'].mean() - ref15
        cw16 = s['LVDT16_Filtered'].mean() - ref16
        # Maximum weighted nominal crack width
        w_max = max(abs(cw14), abs(cw15), abs(cw16))
        print(f"{target:>5} {F:>10.1f} {cw14:>12.6f} {cw15:>12.6f} "
              f"{cw16:>12.6f} {w_max:>14.6f}")
        w_max_list.append(w_max)
        F_list.append(F)

# Shear force at critical section
# Load centroid a = 1500mm from support 5, span L = 10320mm
L = 10.320  # m
a = 1.500   # m

print()
print("SHEAR FORCES AT CRITICAL SECTION:")
V_vals = [F * (L - a) / L for F in F_list]
for i, (F, V) in enumerate(zip(F_list, V_vals)):
    print(f"  F={F:.1f}kN -> V={V:.1f}kN")

print()
print("C++ INPUT VECTORS:")
print(f"std::vector<double> w_max_w = "
      f"{{{', '.join([f'{w:.4f}' for w in w_max_list])}}};")
print(f"std::vector<double> m_Q_PL = "
      f"{{{', '.join([f'{v:.1f}' for v in V_vals])}}};  // kN shear forces")